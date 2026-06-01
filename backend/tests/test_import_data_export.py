from types import SimpleNamespace

from openpyxl import load_workbook

from app.crud.import_data import export_wbs_to_excel


class EmptyQuery:
    def all(self):
        return []


class EmptyDb:
    def query(self, _model):
        return EmptyQuery()


def test_export_wbs_to_excel_handles_numeric_and_empty_values():
    subtask = SimpleNamespace(
        subtask_type_id=None,
        subtask_detail="Subtask",
        ticket_id=123,
        status_id=1,
        assignee_id=None,
        review_days=0.5,
        planned_start_date=None,
        planned_end_date=None,
        planned_effort_days=1.25,
        actual_start_date=None,
        actual_end_date=None,
        actual_effort_days=0.0,
        progress_percent=50,
        workload_percent=100,
        link_url=None,
        memo=None,
    )
    task = SimpleNamespace(
        task_name="Task",
        ticket_id=None,
        status_id=1,
        assignee_id=None,
        planned_start_date=None,
        planned_end_date=None,
        planned_effort_total=1.25,
        actual_start_date=None,
        actual_end_date=None,
        actual_effort_total=0.0,
        progress_percent=50,
        link_url=None,
        memo=None,
        subtasks=[subtask],
    )
    project = SimpleNamespace(
        project_name="Project",
        ticket_id=None,
        status_id=1,
        assignee_id=None,
        planned_start_date=None,
        planned_end_date=None,
        planned_effort_total=1.25,
        actual_start_date=None,
        actual_end_date=None,
        actual_effort_total=0.0,
        progress_percent=50,
        link_url=None,
        memo=None,
        tasks=[task],
    )

    buffer = export_wbs_to_excel([project], EmptyDb())

    workbook = load_workbook(buffer)
    worksheet = workbook["WBS"]
    assert worksheet.max_row == 4
    assert worksheet["A1"].value == "階層"
    assert worksheet["A4"].value == 2
    assert worksheet["K4"].value == 1.25
